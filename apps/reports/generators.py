"""
Generadores de reportes en diferentes formatos.
"""
from django.template.loader import render_to_string
from django.utils import timezone
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
from io import BytesIO
from apps.events.models import Event, Attachment
from apps.catalogs.models import Field, Campaign


class PDFReportGenerator:
    """
    Generador de reportes en formato PDF usando WeasyPrint.
    """
    
    def __init__(self):
        self.font_config = FontConfiguration()
    
    def generate_traceability_report(self, field_id, date_from=None, date_to=None, 
                                    campaign_id=None, event_types=None):
        """
        Genera un reporte PDF de trazabilidad para un lote específico.
        
        Args:
            field_id: ID del lote
            date_from: Fecha de inicio (opcional)
            date_to: Fecha de fin (opcional)
            campaign_id: ID de la campaña (opcional)
            event_types: Lista de tipos de evento a incluir (opcional)
            
        Returns:
            BytesIO: Buffer con el PDF generado
        """
        # Obtener el lote
        try:
            field = Field.objects.get(id=field_id)
        except Field.DoesNotExist:
            raise ValueError(f"Lote con ID {field_id} no encontrado")
        
        # Construir query de eventos
        events_query = Event.objects.filter(field_id=field_id).select_related(
            'event_type', 'campaign', 'created_by'
        )
        
        # Aplicar filtros
        if date_from:
            events_query = events_query.filter(timestamp__gte=date_from)
        if date_to:
            events_query = events_query.filter(timestamp__lte=date_to)
        if campaign_id:
            events_query = events_query.filter(campaign_id=campaign_id)
        if event_types:
            events_query = events_query.filter(event_type_id__in=event_types)
        
        events = events_query.order_by('timestamp')
        
        # Obtener campaña si se especificó
        campaign = None
        if campaign_id:
            try:
                campaign = Campaign.objects.get(id=campaign_id)
            except Campaign.DoesNotExist:
                pass
        
        # Calcular estadísticas
        stats = self._calculate_stats(events)
        
        # Obtener adjuntos
        event_ids = [e.id for e in events]
        attachments = Attachment.objects.filter(event_id__in=event_ids)
        
        # Preparar contexto para el template
        context = {
            'field': field,
            'campaign': campaign,
            'date_from': date_from,
            'date_to': date_to,
            'events': events,
            'stats': stats,
            'attachments': attachments,
            'total_attachments': attachments.count(),
            'generated_at': timezone.now(),
            'report_title': f'Reporte de Trazabilidad - {field.name}',
        }
        
        # Renderizar template HTML
        html_string = render_to_string('reports/pdf/traceability_report.html', context)
        
        # Generar PDF
        html = HTML(string=html_string)
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer, font_config=self.font_config)
        pdf_buffer.seek(0)
        
        return pdf_buffer
    
    def _calculate_stats(self, events):
        """
        Calcula estadísticas del conjunto de eventos.
        
        Args:
            events: QuerySet de eventos
            
        Returns:
            dict: Estadísticas calculadas
        """
        from collections import Counter
        
        # Contar eventos por tipo
        event_type_counts = Counter()
        for event in events:
            event_type_counts[event.event_type.name] += 1
        
        stats = {
            'total': events.count(),
            'by_type': dict(event_type_counts),
        }
        
        return stats
    
    def generate_phytosanitary_report(self, field_id=None, date_from=None, date_to=None):
        """
        Genera un reporte PDF de aplicaciones fitosanitarias.
        
        Args:
            field_id: ID del lote (opcional, si no se especifica incluye todos)
            date_from: Fecha de inicio
            date_to: Fecha de fin
            
        Returns:
            BytesIO: Buffer con el PDF generado
        """
        from apps.events.models import PhytosanitaryEvent
        
        # Construir query
        query = PhytosanitaryEvent.objects.select_related(
            'field', 'campaign', 'created_by'
        )
        
        if field_id:
            query = query.filter(field_id=field_id)
        if date_from:
            query = query.filter(timestamp__gte=date_from)
        if date_to:
            query = query.filter(timestamp__lte=date_to)
        
        applications = query.order_by('timestamp')
        
        # Preparar contexto
        context = {
            'applications': applications,
            'date_from': date_from,
            'date_to': date_to,
            'generated_at': timezone.now(),
            'report_title': 'Registro de Aplicaciones Fitosanitarias',
        }
        
        # Renderizar y generar PDF
        html_string = render_to_string('reports/pdf/phytosanitary_report.html', context)
        html = HTML(string=html_string)
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer, font_config=self.font_config)
        pdf_buffer.seek(0)
        
        return pdf_buffer


    def generate_campaign_traceability_report(self, campaign_id, field_ids=None, event_types=None):
        """
        Genera un reporte PDF de trazabilidad para una campaña completa.
        
        Args:
            campaign_id: ID de la campaña
            field_ids: Lista de IDs de lotes (opcional, todos si no se especifica)
            event_types: Lista de tipos de evento a incluir (opcional)
            
        Returns:
            BytesIO: Buffer con el PDF generado
        """
        # Obtener la campaña
        try:
            campaign = Campaign.objects.get(id=campaign_id)
        except Campaign.DoesNotExist:
            raise ValueError(f"Campaña con ID {campaign_id} no encontrada")
        
        # Construir query de eventos
        events_query = Event.objects.filter(campaign_id=campaign_id).select_related(
            'event_type', 'field', 'created_by'
        )
        
        # Aplicar filtros
        if field_ids:
            events_query = events_query.filter(field_id__in=field_ids)
        if event_types:
            events_query = events_query.filter(event_type_id__in=event_types)
        
        events = events_query.order_by('field__name', 'timestamp')
        
        # Obtener lotes involucrados
        if field_ids:
            fields = Field.objects.filter(id__in=field_ids).order_by('name')
        else:
            # Obtener todos los lotes que tienen eventos en esta campaña
            field_ids_with_events = events.values_list('field_id', flat=True).distinct()
            fields = Field.objects.filter(id__in=field_ids_with_events).order_by('name')
        
        # Calcular estadísticas por lote
        field_stats = []
        for field in fields:
            field_events = events.filter(field=field)
            field_stats.append({
                'field': field,
                'total_events': field_events.count(),
                'by_type': self._calculate_stats(field_events)['by_type'],
                'events': field_events,
            })
        
        # Calcular estadísticas generales
        general_stats = self._calculate_stats(events)
        
        # Preparar contexto para el template
        context = {
            'campaign': campaign,
            'fields': fields,
            'field_stats': field_stats,
            'events': events,
            'general_stats': general_stats,
            'total_fields': fields.count(),
            'generated_at': timezone.now(),
            'report_title': f'Reporte de Trazabilidad - Campaña {campaign.name}',
        }
        
        # Renderizar template HTML
        html_string = render_to_string('reports/pdf/campaign_traceability_report.html', context)
        
        # Generar PDF
        html = HTML(string=html_string)
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer, font_config=self.font_config)
        pdf_buffer.seek(0)
        
        return pdf_buffer


class CSVExporter:
    """
    Exportador de datos a formato CSV.
    """
    
    def export_events(self, field_id=None, date_from=None, date_to=None, 
                     event_types=None):
        """
        Exporta eventos a formato CSV.
        
        Args:
            field_id: ID del lote (opcional)
            date_from: Fecha de inicio (opcional)
            date_to: Fecha de fin (opcional)
            event_types: Lista de tipos de evento (opcional)
            
        Returns:
            list: Lista de diccionarios con los datos
        """
        # Construir query
        query = Event.objects.select_related(
            'event_type', 'field', 'campaign', 'created_by'
        )
        
        if field_id:
            query = query.filter(field_id=field_id)
        if date_from:
            query = query.filter(timestamp__gte=date_from)
        if date_to:
            query = query.filter(timestamp__lte=date_to)
        if event_types:
            query = query.filter(event_type_id__in=event_types)
        
        events = query.order_by('timestamp')
        
        # Convertir a lista de diccionarios
        data = []
        for event in events:
            data.append({
                'id': str(event.id),
                'tipo_evento': event.event_type.name,
                'lote': event.field.name,
                'campana': event.campaign.name if event.campaign else '',
                'fecha_hora': event.timestamp.strftime('%Y-%m-%d %H:%M'),
                'observaciones': event.observations or '',
                'creado_por': event.created_by.get_full_name() if event.created_by else '',
                'creado_el': event.created_at.strftime('%Y-%m-%d %H:%M'),
            })
        
        return data


    def export_campaign_events(self, campaign_id, field_ids=None, event_types=None):
        """
        Exporta eventos de una campaña a formato CSV.
        
        Args:
            campaign_id: ID de la campaña
            field_ids: Lista de IDs de lotes (opcional)
            event_types: Lista de tipos de evento (opcional)
            
        Returns:
            list: Lista de diccionarios con los datos
        """
        # Construir query
        query = Event.objects.filter(campaign_id=campaign_id).select_related(
            'event_type', 'field', 'campaign', 'created_by'
        )
        
        if field_ids:
            query = query.filter(field_id__in=field_ids)
        if event_types:
            query = query.filter(event_type_id__in=event_types)
        
        events = query.order_by('field__name', 'timestamp')
        
        # Convertir a lista de diccionarios
        data = []
        for event in events:
            data.append({
                'id': str(event.id),
                'campana': event.campaign.name if event.campaign else '',
                'lote': event.field.name,
                'tipo_evento': event.event_type.name,
                'categoria': event.event_type.category,
                'fecha_hora': event.timestamp.strftime('%Y-%m-%d %H:%M'),
                'observaciones': event.observations or '',
                'creado_por': event.created_by.get_full_name() if event.created_by else '',
                'creado_el': event.created_at.strftime('%Y-%m-%d %H:%M'),
            })
        
        return data


class ExcelExporter:
    """
    Exportador de datos a formato Excel (XLSX).
    """
    
    def export_events(self, field_id=None, date_from=None, date_to=None, 
                     event_types=None):
        """
        Exporta eventos a formato Excel con múltiples hojas.
        
        Args:
            field_id: ID del lote (opcional)
            date_from: Fecha de inicio (opcional)
            date_to: Fecha de fin (opcional)
            event_types: Lista de tipos de evento (opcional)
            
        Returns:
            BytesIO: Buffer con el archivo Excel generado
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Eventos
        ws_events = wb.active
        ws_events.title = "Eventos"
        
        # Encabezados
        headers = ['ID', 'Tipo Evento', 'Lote', 'Campaña', 'Fecha/Hora', 
                  'Observaciones', 'Creado Por', 'Creado El']
        
        # Aplicar estilo a encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws_events.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Obtener datos usando CSVExporter
        csv_exporter = CSVExporter()
        data = csv_exporter.export_events(field_id, date_from, date_to, event_types)
        
        # Escribir datos
        for row_idx, event_data in enumerate(data, 2):
            ws_events.cell(row=row_idx, column=1, value=event_data['id'])
            ws_events.cell(row=row_idx, column=2, value=event_data['tipo_evento'])
            ws_events.cell(row=row_idx, column=3, value=event_data['lote'])
            ws_events.cell(row=row_idx, column=4, value=event_data['campana'])
            ws_events.cell(row=row_idx, column=5, value=event_data['fecha_hora'])
            ws_events.cell(row=row_idx, column=6, value=event_data['observaciones'])
            ws_events.cell(row=row_idx, column=7, value=event_data['creado_por'])
            ws_events.cell(row=row_idx, column=8, value=event_data['creado_el'])
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws_events.column_dimensions[ws_events.cell(row=1, column=col).column_letter].width = 20
        
        # Hoja 2: Resumen
        ws_summary = wb.create_sheet("Resumen")
        ws_summary.cell(row=1, column=1, value="Resumen de Exportación").font = Font(bold=True, size=14)
        ws_summary.cell(row=3, column=1, value="Total de eventos:")
        ws_summary.cell(row=3, column=2, value=len(data))
        ws_summary.cell(row=4, column=1, value="Fecha de generación:")
        ws_summary.cell(row=4, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M'))
        
        # Guardar en buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer


    def export_campaign_events(self, campaign_id, field_ids=None, event_types=None):
        """
        Exporta eventos de una campaña a formato Excel con múltiples hojas.
        
        Args:
            campaign_id: ID de la campaña
            field_ids: Lista de IDs de lotes (opcional)
            event_types: Lista de tipos de evento (opcional)
            
        Returns:
            BytesIO: Buffer con el archivo Excel generado
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Crear workbook
        wb = Workbook()
        
        # Obtener campaña
        campaign = Campaign.objects.get(id=campaign_id)
        
        # Hoja 1: Eventos
        ws_events = wb.active
        ws_events.title = "Eventos"
        
        # Encabezados
        headers = ['ID', 'Campaña', 'Lote', 'Tipo Evento', 'Categoría', 
                  'Fecha/Hora', 'Observaciones', 'Creado Por', 'Creado El']
        
        # Aplicar estilo a encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws_events.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Obtener datos usando CSVExporter
        csv_exporter = CSVExporter()
        data = csv_exporter.export_campaign_events(campaign_id, field_ids, event_types)
        
        # Escribir datos
        for row_idx, event_data in enumerate(data, 2):
            ws_events.cell(row=row_idx, column=1, value=event_data['id'])
            ws_events.cell(row=row_idx, column=2, value=event_data['campana'])
            ws_events.cell(row=row_idx, column=3, value=event_data['lote'])
            ws_events.cell(row=row_idx, column=4, value=event_data['tipo_evento'])
            ws_events.cell(row=row_idx, column=5, value=event_data['categoria'])
            ws_events.cell(row=row_idx, column=6, value=event_data['fecha_hora'])
            ws_events.cell(row=row_idx, column=7, value=event_data['observaciones'])
            ws_events.cell(row=row_idx, column=8, value=event_data['creado_por'])
            ws_events.cell(row=row_idx, column=9, value=event_data['creado_el'])
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws_events.column_dimensions[ws_events.cell(row=1, column=col).column_letter].width = 18
        
        # Hoja 2: Resumen por Lote
        ws_summary = wb.create_sheet("Resumen por Lote")
        ws_summary.cell(row=1, column=1, value="Resumen de Eventos por Lote").font = Font(bold=True, size=14)
        
        # Calcular estadísticas por lote
        from collections import Counter
        field_counter = Counter()
        for event in data:
            field_counter[event['lote']] += 1
        
        ws_summary.cell(row=3, column=1, value="Lote").font = Font(bold=True)
        ws_summary.cell(row=3, column=2, value="Total Eventos").font = Font(bold=True)
        
        row = 4
        for field_name, count in sorted(field_counter.items()):
            ws_summary.cell(row=row, column=1, value=field_name)
            ws_summary.cell(row=row, column=2, value=count)
            row += 1
        
        # Hoja 3: Información de la Campaña
        ws_info = wb.create_sheet("Información")
        ws_info.cell(row=1, column=1, value="Información de la Campaña").font = Font(bold=True, size=14)
        ws_info.cell(row=3, column=1, value="Campaña:")
        ws_info.cell(row=3, column=2, value=campaign.name)
        ws_info.cell(row=4, column=1, value="Fecha Inicio:")
        ws_info.cell(row=4, column=2, value=campaign.start_date.strftime('%Y-%m-%d') if campaign.start_date else '')
        ws_info.cell(row=5, column=1, value="Fecha Fin:")
        ws_info.cell(row=5, column=2, value=campaign.end_date.strftime('%Y-%m-%d') if campaign.end_date else 'Activa')
        ws_info.cell(row=6, column=1, value="Estado:")
        ws_info.cell(row=6, column=2, value='Activa' if campaign.is_active else 'Finalizada')
        ws_info.cell(row=7, column=1, value="Total de eventos:")
        ws_info.cell(row=7, column=2, value=len(data))
        ws_info.cell(row=8, column=1, value="Fecha de generación:")
        ws_info.cell(row=8, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M'))
        
        # Guardar en buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
